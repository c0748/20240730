import openpyxl

wb = openpyxl.load_workbook("files/new_wb.xlsx")

# # ws = wb.active

# ws = wb["Sheet"]
# ws.title = "Sheet3"

ws = wb.active

ws["A1"].value = 5
ws["B1"].value = 10
ws["C1"].value = "=B1*A1"


print(ws["C1"].value)

wb.save("files/new_wb.xlsx")

# print(ws)