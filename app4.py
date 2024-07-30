import openpyxl

wb = openpyxl.Workbook()

ws = wb.active

ws.merge_cells("B1:D1")

ws["B1"].value = "Merge Cell Value"


print(ws["C1"])



wb.save("files/merge.xlsx")
