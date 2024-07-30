import openpyxl

wb = openpyxl.Workbook()
wb.save("files/new_wb.xlsx")

sheet2 = wb.create_sheet(title="sheet2")

wb.create_sheet(title="sheet3")

wb.save("files/new_wb2.xlsx")

wb.remove(sheet2)

ws = wb["sheet3"]
wb.copy_worksheet(ws)

wb.save("files/new_wb2.xlsx")
